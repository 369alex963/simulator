"""CSV, XLSX, and PDF export endpoints."""
import csv
import io
import re
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from apps.accounts.models import Role
from apps.accounts.permissions import IsTeacherOrAbove
from apps.enrollments.models import Enrollment, QuestionAttempt
from apps.instances.models import Instance
from apps.instances.views import _instance_access


def _safe_filename(name: str) -> str:
    """Sanitize a string for use in a Content-Disposition filename header."""
    return re.sub(r'[^\w\-]', '_', name)[:60]


def _csv_safe(value) -> str:
    """Escape CSV formula injection (=, +, -, @, TAB, CR at start of cell)."""
    s = str(value) if value is not None else ""
    if s and s[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + s
    return s


def _get_instance_or_403(request, pk):
    try:
        instance = Instance.objects.select_related("scenario", "branch").get(pk=pk)
    except Instance.DoesNotExist:
        return None, Response(status=404)
    if not _instance_access(request.user, instance):
        return None, Response(status=403)
    return instance, None


@api_view(["GET"])
@permission_classes([IsTeacherOrAbove])
def export_csv(request, pk):
    """GET /api/exports/<pk>/csv/?mode=summary|analytics|full"""
    instance, err = _get_instance_or_403(request, pk)
    if err:
        return err

    mode = request.query_params.get("mode", "summary")
    buf = io.StringIO()

    if mode in ("summary", "full"):
        _write_summary(buf, instance)
    if mode in ("analytics", "full"):
        if mode == "full":
            buf.write("\n\n")
        _write_analytics(buf, instance)

    instance.exported_at = timezone.now()
    instance.exported_by = request.user
    instance.save(update_fields=["exported_at", "exported_by"])

    safe_name = _safe_filename(instance.name)
    response = HttpResponse(buf.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="{safe_name}-{mode}-{timezone.now().date()}.csv"'
    )
    return response


def _fmt_duration(total_seconds: int) -> str:
    """Format seconds as H:MM:SS (or M:SS for short durations)."""
    if total_seconds is None:
        return ""
    total = int(total_seconds)
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


def _write_summary(buf, instance):
    writer = csv.writer(buf)
    writer.writerow([
        "Student", "Email", "Status", "Score",
        "Started", "Submitted", "Completed",
        "Wall Clock (s)", "Active Time (s)",
    ])
    for e in instance.enrollments.select_related("user").prefetch_related("attempts").order_by("-score_total"):
        wall = ""
        if e.started_at:
            end = e.submitted_at or e.completed_at
            if end:
                wall = int((end - e.started_at).total_seconds())
        active = sum(a.active_seconds for a in e.attempts.all())
        writer.writerow([
            _csv_safe(e.user.username),
            _csv_safe(e.user.email),
            _csv_safe(e.status),
            float(e.score_total),
            e.started_at.isoformat() if e.started_at else "",
            e.submitted_at.isoformat() if e.submitted_at else "",
            e.completed_at.isoformat() if e.completed_at else "",
            wall,
            active,
        ])


def _write_analytics(buf, instance):
    questions = list(instance.scenario.questions.order_by("order")) if instance.scenario else []
    writer = csv.writer(buf)
    writer.writerow(["Order", "Title", "Type", "Bonus", "Base Pts", "Total Attempts",
                     "Correct", "Pct Correct", "Avg Attempts", "Avg Time (s)"])
    for q in questions:
        qs = QuestionAttempt.objects.filter(enrollment__instance=instance, question=q)
        total = qs.count()
        correct = qs.filter(is_correct=True).count()
        from django.db.models import Avg
        avg_att = float(qs.aggregate(v=Avg("attempts"))["v"] or 0)
        avg_t = float(qs.aggregate(v=Avg("active_seconds"))["v"] or 0)
        writer.writerow([
            q.order,
            _csv_safe(q.title),
            _csv_safe(q.question_type),
            q.is_bonus,
            float(q.base_points),
            total, correct,
            round(correct / total * 100, 1) if total else 0,
            round(avg_att, 2), round(avg_t, 1),
        ])


@api_view(["GET"])
@permission_classes([IsTeacherOrAbove])
def export_xlsx(request, pk):
    """GET /api/exports/<pk>/xlsx/"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response({"detail": "openpyxl not installed."}, status=501)

    instance, err = _get_instance_or_403(request, pk)
    if err:
        return err

    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    gold = "FFD700"
    dark = "0D0F14"
    header_font = Font(bold=True, color=dark)
    header_fill = PatternFill("solid", fgColor=gold)

    headers = [
        "Student", "Email", "Status", "Score",
        "Started", "Submitted", "Completed",
        "Wall Clock (s)", "Active Time (s)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws1.freeze_panes = "A2"

    for e in instance.enrollments.select_related("user").prefetch_related("attempts").order_by("-score_total"):
        wall = ""
        if e.started_at:
            end = e.submitted_at or e.completed_at
            if end:
                wall = int((end - e.started_at).total_seconds())
        active = sum(a.active_seconds for a in e.attempts.all())
        ws1.append([
            _csv_safe(e.user.username),
            _csv_safe(e.user.email),
            _csv_safe(e.status),
            float(e.score_total),
            e.started_at.isoformat() if e.started_at else "",
            e.submitted_at.isoformat() if e.submitted_at else "",
            e.completed_at.isoformat() if e.completed_at else "",
            wall,
            active,
        ])

    # ── Analytics sheet ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Analytics")
    ana_headers = ["Order", "Title", "Type", "Bonus", "Base Pts",
                   "Total Attempts", "Correct", "% Correct", "Avg Attempts", "Avg Time (s)"]
    for col, h in enumerate(ana_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    ws2.freeze_panes = "A2"

    from django.db.models import Avg
    questions = instance.scenario.questions.order_by("order") if instance.scenario else []
    for q in questions:
        qs = QuestionAttempt.objects.filter(enrollment__instance=instance, question=q)
        total = qs.count()
        correct = qs.filter(is_correct=True).count()
        ws2.append([
            q.order,
            _csv_safe(q.title),
            _csv_safe(q.question_type),
            q.is_bonus,
            float(q.base_points),
            total, correct,
            round(correct / total * 100, 1) if total else 0,
            float(qs.aggregate(v=Avg("attempts"))["v"] or 0),
            float(qs.aggregate(v=Avg("active_seconds"))["v"] or 0),
        ])

    instance.exported_at = timezone.now()
    instance.exported_by = request.user
    instance.save(update_fields=["exported_at", "exported_by"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = _safe_filename(instance.name)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{safe_name}-results-{timezone.now().date()}.xlsx"'
    )
    return response


@api_view(["GET"])
@permission_classes([IsTeacherOrAbove])
def export_pdf(request, pk):
    """GET /api/exports/<pk>/pdf/?mode=summary|analytics|full"""
    instance, err = _get_instance_or_403(request, pk)
    if err:
        return err

    mode = request.query_params.get("mode", "full")

    from .pdf import build_pdf
    pdf_bytes = build_pdf(instance, mode, exporter=request.user)
    if not pdf_bytes:
        return Response({"detail": "reportlab not installed."}, status=501)

    instance.exported_at = timezone.now()
    instance.exported_by = request.user
    instance.save(update_fields=["exported_at", "exported_by"])

    safe_name = _safe_filename(instance.name)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="{safe_name}-{mode}-{timezone.now().date()}.pdf"'
    )
    return response


@api_view(["POST"])
@permission_classes([IsTeacherOrAbove])
def recalculate_scores(request, pk):
    """POST /api/exports/<pk>/recalculate/ — recompute all enrollment scores."""
    instance, err = _get_instance_or_403(request, pk)
    if err:
        return err

    from apps.enrollments.scoring import update_enrollment_score

    updated = 0
    for e in instance.enrollments.all():
        update_enrollment_score(e)
        updated += 1

    return Response({"recalculated": updated, "instance": instance.name})
